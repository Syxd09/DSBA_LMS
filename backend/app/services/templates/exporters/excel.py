"""
EduMetrics Templates - Excel Exporter

Generates Excel reports from ReportOutput using openpyxl.

PHASE-2C CONSTRAINT: Export only, no computation.
"""
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional, Dict, Any
from decimal import Decimal

# Lazy import to avoid dependency issues if openpyxl not installed
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Border, Side,
        Alignment, NamedStyle
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.services.templates.base import (
    ReportOutput,
    ReportSection,
    Table,
    TableRow,
    TableCell,
)


class ExcelExporter:
    """
    Excel Exporter for ReportOutput.
    
    Uses openpyxl to generate .xlsx files.
    """
    
    def __init__(self):
        """Initialize Excel exporter."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )
        
        # Define styles
        self._init_styles()
    
    def _init_styles(self):
        """Initialize cell styles."""
        self.header_font = Font(bold=True, size=10)
        self.title_font = Font(bold=True, size=14)
        self.section_font = Font(bold=True, size=12)
        
        self.header_fill = PatternFill(
            start_color="E0E0E0",
            end_color="E0E0E0",
            fill_type="solid"
        )
        
        self.level_fills = {
            0: PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
            1: PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
            2: PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),
            3: PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid"),
        }
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.wrap_align = Alignment(wrap_text=True, vertical='top')
    
    def export(self, report: ReportOutput) -> bytes:
        """
        Export ReportOutput to Excel bytes.
        
        Args:
            report: ReportOutput object to export
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        
        # Main sheet
        ws = wb.active
        ws.title = "Report"
        
        current_row = 1
        
        # Header
        current_row = self._write_header(ws, report, current_row)
        current_row += 2
        
        # Sections
        for section in report.sections:
            current_row = self._write_section(ws, section, current_row)
            current_row += 2
        
        # Warnings
        if report.warnings:
            current_row = self._write_warnings(ws, report.warnings, current_row)
        
        # Footer
        self._write_footer(ws, report, current_row + 2)
        
        # Auto-fit columns (approximate)
        self._auto_fit_columns(ws)
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def export_to_file(self, report: ReportOutput, path: Path) -> None:
        """Export ReportOutput to Excel file."""
        excel_bytes = self.export(report)
        path.write_bytes(excel_bytes)
    
    def _write_header(self, ws, report: ReportOutput, row: int) -> int:
        """Write report header."""
        m = report.metadata
        
        # Institution
        ws.cell(row=row, column=1, value=m.institution_name or "Institution")
        ws.cell(row=row, column=1).font = self.title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        
        # Department
        ws.cell(row=row, column=1, value=m.department_name)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        
        # Title
        ws.cell(row=row, column=1, value=report.title)
        ws.cell(row=row, column=1).font = self.section_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        
        # Metadata row
        meta_parts = []
        if m.program_name:
            meta_parts.append(f"Program: {m.program_name}")
        if m.batch:
            meta_parts.append(f"Batch: {m.batch}")
        if m.academic_year:
            meta_parts.append(f"Year: {m.academic_year}")
        if m.semester:
            meta_parts.append(f"Semester: {m.semester}")
        
        ws.cell(row=row, column=1, value=" | ".join(meta_parts))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        
        return row
    
    def _write_section(self, ws, section: ReportSection, row: int) -> int:
        """Write a report section."""
        # Section title
        ws.cell(row=row, column=1, value=section.title)
        ws.cell(row=row, column=1).font = self.section_font
        row += 1
        
        # Content
        if isinstance(section.content, Table):
            row = self._write_table(ws, section.content, row)
        elif isinstance(section.content, dict):
            row = self._write_dict(ws, section.content, row)
        else:
            ws.cell(row=row, column=1, value=str(section.content))
            row += 1
        
        # Notes
        if section.notes:
            row += 1
            for note in section.notes:
                ws.cell(row=row, column=1, value=f"• {note}")
                ws.cell(row=row, column=1).font = Font(italic=True, size=9)
                row += 1
        
        # Subsections
        for subsection in section.subsections:
            row = self._write_section(ws, subsection, row + 1)
        
        return row
    
    def _write_table(self, ws, table: Table, row: int) -> int:
        """Write a table."""
        start_col = 1
        
        # Headers
        for col_idx, header in enumerate(table.headers, start=start_col):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_align
        row += 1
        
        # Data rows
        for table_row in table.rows:
            for col_idx, cell_data in enumerate(table_row.cells, start=start_col):
                value = cell_data.value
                if isinstance(value, Decimal):
                    value = float(value)
                
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = self.thin_border
                
                # Apply level styling
                if "level-" in cell_data.css_class:
                    level = int(cell_data.css_class.split("-")[1][0])
                    if level in self.level_fills:
                        cell.fill = self.level_fills[level]
            row += 1
        
        # Caption
        if table.caption:
            ws.cell(row=row, column=1, value=table.caption)
            ws.cell(row=row, column=1).font = Font(italic=True, size=9)
            row += 1
        
        return row
    
    def _write_dict(self, ws, data: dict, row: int) -> int:
        """Write dictionary as key-value pairs."""
        for key, value in data.items():
            if key.startswith("_"):
                continue
            
            formatted_key = key.replace("_", " ").title()
            
            ws.cell(row=row, column=1, value=formatted_key)
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            ws.cell(row=row, column=2, value=str(value))
            
            row += 1
        
        return row
    
    def _write_warnings(self, ws, warnings: list, row: int) -> int:
        """Write warnings section."""
        ws.cell(row=row, column=1, value="Data Warnings")
        ws.cell(row=row, column=1).font = self.section_font
        row += 1
        
        for w in warnings:
            msg = w.get("message", str(w))
            ws.cell(row=row, column=1, value=f"⚠ {msg}")
            ws.cell(row=row, column=1).font = Font(color="FF0000", size=9)
            row += 1
        
        return row
    
    def _write_footer(self, ws, report: ReportOutput, row: int) -> None:
        """Write report footer."""
        m = report.metadata
        
        footer_text = (
            f"Generated: {m.generated_at.strftime('%Y-%m-%d %H:%M:%S')} UTC | "
            f"By: {m.generated_by} | Source: {m.source_version}"
        )
        
        ws.cell(row=row, column=1, value=footer_text)
        ws.cell(row=row, column=1).font = Font(italic=True, size=9, color="666666")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    
    def _auto_fit_columns(self, ws, min_width: int = 10, max_width: int = 50) -> None:
        """Auto-fit column widths (approximate)."""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max(min_width, max_length + 2), max_width)
            ws.column_dimensions[column].width = adjusted_width
