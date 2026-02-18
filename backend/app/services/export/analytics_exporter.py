"""
EduMetrics - Universal Analytics Exporter

Provides unified export for analytics data in multiple formats:
- JSON (structured data)
- CSV (tabular data)
- XLSX (styled spreadsheet)
- PDF (formatted report)

Works with raw dict data from analytics endpoints.
"""
from datetime import datetime
from enum import Enum
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Union
import csv
import json

from fastapi.responses import Response, StreamingResponse


class ExportFormat(str, Enum):
    JSON = "json"
    CSV = "csv"
    XLSX = "xlsx"
    PDF = "pdf"


class AnalyticsExporter:
    """
    Universal exporter for analytics data.
    
    Handles different data shapes:
    - Single records (dict)
    - Lists of records 
    - Nested data with flattening
    """
    
    @staticmethod
    def export(
        data: Union[Dict, List],
        format: ExportFormat,
        filename: str,
        title: str = "",
        headers: Optional[List[str]] = None,
        flatten_key: Optional[str] = None
    ) -> Response:
        """
        Export analytics data in requested format.
        
        Args:
            data: Dict or list to export
            format: Export format (json, csv, xlsx, pdf)
            filename: Base filename without extension
            title: Report title (for PDF/XLSX)
            headers: Column headers for CSV/XLSX (auto-detected if None)
            flatten_key: Key containing list to flatten for tabular export
        
        Returns:
            FastAPI Response with appropriate content type
        """
        if format == ExportFormat.JSON:
            return AnalyticsExporter._export_json(data, filename)
        elif format == ExportFormat.CSV:
            return AnalyticsExporter._export_csv(data, filename, headers, flatten_key)
        elif format == ExportFormat.XLSX:
            return AnalyticsExporter._export_xlsx(data, filename, title, headers, flatten_key)
        elif format == ExportFormat.PDF:
            return AnalyticsExporter._export_pdf(data, filename, title)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    @staticmethod
    def _export_json(data: Union[Dict, List], filename: str) -> Response:
        """Export as JSON."""
        content = json.dumps(data, indent=2, default=str)
        return Response(
            content=content,
            media_type="application/json",
            headers={
                "Content-Disposition": f"attachment; filename={filename}.json"
            }
        )
    
    @staticmethod
    def _export_csv(
        data: Union[Dict, List], 
        filename: str,
        headers: Optional[List[str]] = None,
        flatten_key: Optional[str] = None
    ) -> Response:
        """Export as CSV."""
        output = StringIO()
        
        # Handle different data shapes
        if isinstance(data, dict):
            if flatten_key and flatten_key in data:
                rows = data[flatten_key]
            else:
                # Convert single dict to list of key-value pairs
                rows = [{"key": k, "value": str(v)} for k, v in data.items() 
                        if not isinstance(v, (dict, list))]
        else:
            rows = data
        
        if not rows:
            output.write("No data available\n")
        else:
            # Auto-detect headers from first row
            if headers is None and rows:
                first_row = rows[0]
                if isinstance(first_row, dict):
                    headers = list(first_row.keys())
                else:
                    headers = ["value"]
            
            writer = csv.DictWriter(output, fieldnames=headers, extrasaction='ignore')
            writer.writeheader()
            
            for row in rows:
                if isinstance(row, dict):
                    # Flatten nested dicts
                    flat_row = {}
                    for k, v in row.items():
                        if isinstance(v, (dict, list)):
                            flat_row[k] = json.dumps(v)
                        else:
                            flat_row[k] = v
                    writer.writerow(flat_row)
                else:
                    writer.writerow({"value": row})
        
        content = output.getvalue()
        return Response(
            content=content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}.csv"
            }
        )
    
    @staticmethod
    def _export_xlsx(
        data: Union[Dict, List],
        filename: str,
        title: str = "",
        headers: Optional[List[str]] = None,
        flatten_key: Optional[str] = None
    ) -> Response:
        """Export as styled Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl required for Excel export. Install: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytics Report"
        
        # Styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row_num = 1
        
        # Title row
        if title:
            ws.cell(row=row_num, column=1, value=title)
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
            row_num += 2
        
        # Get rows
        if isinstance(data, dict):
            if flatten_key and flatten_key in data:
                rows = data[flatten_key]
                # Add summary info first
                for k, v in data.items():
                    if k != flatten_key and not isinstance(v, (dict, list)):
                        ws.cell(row=row_num, column=1, value=str(k))
                        ws.cell(row=row_num, column=2, value=str(v))
                        row_num += 1
                row_num += 1
            else:
                rows = [{"key": k, "value": str(v)} for k, v in data.items() 
                        if not isinstance(v, (dict, list))]
        else:
            rows = data
        
        if rows:
            # Header row
            if headers is None:
                first_row = rows[0]
                if isinstance(first_row, dict):
                    headers = list(first_row.keys())
                else:
                    headers = ["value"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header.replace("_", " ").title())
                cell.font = header_font
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = header_alignment
                cell.border = thin_border
            row_num += 1
            
            # Data rows
            for row in rows:
                for col, header in enumerate(headers, 1):
                    if isinstance(row, dict):
                        value = row.get(header, "")
                        if isinstance(value, (dict, list)):
                            value = json.dumps(value)
                    else:
                        value = row
                    
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = thin_border
                row_num += 1
            
            # Auto-fit columns
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                max_length = max(
                    len(str(ws.cell(row=r, column=col).value or ""))
                    for r in range(1, row_num)
                )
                ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)
        
        # Footer
        row_num += 1
        ws.cell(row=row_num, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=row_num, column=1).font = Font(italic=True, size=9)
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}.xlsx"
            }
        )
    
    @staticmethod
    def _export_pdf(
        data: Union[Dict, List],
        filename: str,
        title: str = ""
    ) -> Response:
        """Export as PDF report."""
        try:
            from weasyprint import HTML, CSS
        except ImportError:
            raise ImportError("weasyprint required for PDF export. Install: pip install weasyprint")
        
        # Build HTML
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 40px;
                    font-size: 12px;
                }}
                h1 {{
                    color: #2c3e50;
                    border-bottom: 2px solid #3498db;
                    padding-bottom: 10px;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 20px;
                }}
                th {{
                    background-color: #3498db;
                    color: white;
                    padding: 10px;
                    text-align: left;
                }}
                td {{
                    padding: 8px;
                    border-bottom: 1px solid #ddd;
                }}
                tr:nth-child(even) {{
                    background-color: #f9f9f9;
                }}
                .summary {{
                    background-color: #ecf0f1;
                    padding: 15px;
                    border-radius: 5px;
                    margin-bottom: 20px;
                }}
                .footer {{
                    margin-top: 30px;
                    font-size: 10px;
                    color: #7f8c8d;
                }}
            </style>
        </head>
        <body>
            <h1>{title or 'Analytics Report'}</h1>
        """
        
        # Add summary if dict with non-list values
        if isinstance(data, dict):
            summary_items = [(k, v) for k, v in data.items() 
                           if not isinstance(v, (dict, list))]
            if summary_items:
                html_content += '<div class="summary">'
                for k, v in summary_items:
                    html_content += f'<p><strong>{k.replace("_", " ").title()}:</strong> {v}</p>'
                html_content += '</div>'
            
            # Find list to render as table
            for key, value in data.items():
                if isinstance(value, list) and value:
                    html_content += AnalyticsExporter._list_to_html_table(value, key.replace("_", " ").title())
        elif isinstance(data, list):
            html_content += AnalyticsExporter._list_to_html_table(data, "Data")
        
        html_content += f"""
            <div class="footer">
                Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
            </div>
        </body>
        </html>
        """
        
        # Generate PDF
        pdf = HTML(string=html_content).write_pdf()
        
        return Response(
            content=pdf,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}.pdf"
            }
        )
    
    @staticmethod
    def _list_to_html_table(data: List, title: str = "") -> str:
        """Convert list of dicts to HTML table."""
        if not data:
            return "<p>No data</p>"
        
        first_row = data[0]
        if not isinstance(first_row, dict):
            return f"<p>{', '.join(str(x) for x in data)}</p>"
        
        headers = list(first_row.keys())
        
        html = f"<h3>{title}</h3><table><thead><tr>"
        for header in headers:
            html += f"<th>{header.replace('_', ' ').title()}</th>"
        html += "</tr></thead><tbody>"
        
        for row in data:
            html += "<tr>"
            for header in headers:
                value = row.get(header, "")
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                html += f"<td>{value}</td>"
            html += "</tr>"
        
        html += "</tbody></table>"
        return html


# Convenience function for analytics endpoints
def export_analytics(
    data: Union[Dict, List],
    format: ExportFormat,
    filename: str,
    title: str = "",
    headers: Optional[List[str]] = None,
    flatten_key: Optional[str] = None
) -> Response:
    """
    Export analytics data.
    
    Usage in endpoint:
        @router.get("/analytics/data")
        async def get_data(format: ExportFormat = Query(default=ExportFormat.JSON)):
            data = await fetch_analytics()
            return export_analytics(data, format, "my_report", "My Report Title")
    """
    return AnalyticsExporter.export(
        data=data,
        format=format,
        filename=filename,
        title=title,
        headers=headers,
        flatten_key=flatten_key
    )
