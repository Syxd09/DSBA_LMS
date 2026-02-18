"""
EduMetrics Backend - Marks Router
Marks entry and management endpoints

NOTE: MarksComputed has been REMOVED per the "no stored analytics" rule.
All computation is now on-demand via Phase-2A functions.
"""
from typing import List
from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from sqlalchemy import func
from uuid import UUID
import uuid as uuid_lib
from datetime import datetime
from decimal import Decimal

from app.database import get_db
from app.api.deps import get_current_user, require_teacher_or_above, require_authenticated
from app.models import Profile, Exam, StudentQuestionMark, SubQuestion, Question, ExamSection, Student
from app.schemas import BulkMarksCreate, StudentMarksResponse
from app.core.cache import cache_manager, settings  # Import cache

router = APIRouter(prefix="/marks", tags=["Marks"])



from typing import Optional
from fastapi import BackgroundTasks

def invalidate_analytics_cache_bg(offering_id: UUID, program_id: Optional[UUID]):
    """
    Background task to invalidate analytics cache.
    Does NOT require DB session (pure Redis).
    """
    try:
        # Invalidate CO cache
        cache_manager.delete(f"co_attainment:{offering_id}")
        
        if program_id:
            from app.core.cache import invalidate_cache
            invalidate_cache(f"po_attainment:{program_id}:*")
            
    except Exception as e:
        print(f"Background cache invalidation failed: {e}")


def verify_marks_entry_access(db: Session, user: Profile, exam: Exam):
    """
    Verify if user is authorized to enter marks for this exam.
    Allowed:
    - Creator of the exam
    - Teacher assigned to the offering
    - HOD/Principal/Admin
    """
    # 1. Check Roles (HOD/Principal can always edit if status allows)
    # Note: Profile might not have user_role loaded if using simple dependency, 
    # but require_teacher_or_above usually implies role check.
    # We should query UserRole explicitly to be safe or rely on joined load.
    from app.models import UserRole
    from app.core.permissions import AppRole
    
    role_entry = db.query(UserRole).filter(UserRole.user_id == user.user_id).first()
    if role_entry and role_entry.role == AppRole.ADMIN:
        return True
    
    if role_entry and role_entry.role == AppRole.PRINCIPAL:
        return True

    # [NEW] Creator always has access
    if exam.teacher_id == user.user_id:
        return True

    if role_entry and role_entry.role == AppRole.HOD:
        from app.models import Department, Program, Subject
        from app.models.subject_offering import SubjectOffering
        # Get HOD's department
        dept = db.query(Department).filter(Department.hod_id == user.user_id).first()
        if not dept:
            raise HTTPException(status_code=403, detail="HOD record not found for this user")
            
        # Get Exam's department
        offering = None
        if exam.offering_id:
            offering = db.query(SubjectOffering).filter(SubjectOffering.id == exam.offering_id).first()
        
        program_id = None
        if offering:
            program_id = offering.program_id
        elif exam.subject_id:
            # Fallback to subject -> curriculum -> program
            from app.models.academic import CurriculumVersion
            subject = db.query(Subject).filter(Subject.id == exam.subject_id).first()
            if subject and subject.curriculum_version_id:
                cv = db.query(CurriculumVersion).filter(CurriculumVersion.id == subject.curriculum_version_id).first()
                if cv:
                    program_id = cv.program_id
        
        if not program_id:
             raise HTTPException(status_code=404, detail="Exam metadata (offering/subject) is missing or orphaned")
        
        program = db.query(Program).filter(Program.id == program_id).first()
        if not program:
             raise HTTPException(status_code=404, detail="Program not found for exam")

        if program.department_id != dept.id:
            raise HTTPException(
                status_code=403, 
                detail=f"HOD access denied. Exam belongs to department {program.department_id}, but user is HOD of {dept.id}"
            )
        return True
        
    # 3. Check Assignment
    if exam.offering_id:
        from app.models.academic import TeacherAssignment
        assignment = db.query(TeacherAssignment).filter(
            TeacherAssignment.teacher_id == user.user_id,
            TeacherAssignment.offering_id == exam.offering_id
        ).first()
        if assignment:
            return True
            
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN, 
        detail="Not authorized to enter marks for this exam (Not creator or assigned teacher)"
    )



@router.get("/exam/{exam_id}", response_model=List[StudentMarksResponse])
def get_exam_marks(
    exam_id: UUID,
    db: Session = Depends(get_db),
    current_user: Profile = Depends(require_teacher_or_above)
):
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
        
    # [RBAC] Verify access
    verify_marks_entry_access(db, current_user, exam)
    
    marks = db.query(StudentQuestionMark).filter(StudentQuestionMark.exam_id == exam_id).all()
    
    return [
        {
            "id": m.id,
            "exam_id": m.exam_id,
            "student_id": m.usn, # Return USN directly
            "sub_question_id": m.sub_question_id,
            "marks": m.marks,
            "entered_at": m.entered_at
        }
        for m in marks
    ]


@router.post("/exam/{exam_id}", response_model=dict)
def save_marks(
    exam_id: UUID,
    marks_data: BulkMarksCreate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: Profile = Depends(require_teacher_or_above)
):
    """
    Bulk save marks for an exam.
    
    IMMUTABILITY RULES:
    - Only 'approved' status allows marks entry
    - 'draft' and 'submitted' block marks (exam structure not finalized)
    - 'locked' blocks all edits (immutable)
    
    AUDIT LOGGING:
    - Every mark entry/edit is logged with old/new values
    """
    from app.models import AuditLog
    
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    # Immutability enforcement (CRITICAL)
    if exam.status == "locked":
        raise HTTPException(
            status_code=400, 
            detail="Exam is locked. Marks cannot be modified. Contact HOD to unlock."
        )
    
    if exam.status in ["draft", "submitted"]:
        raise HTTPException(
            status_code=400,
            detail=f"Exam status is '{exam.status}'. Marks entry only allowed after HOD approval."
        )
    
    if exam.status != "approved":
        raise HTTPException(
            status_code=400,
            detail=f"Unexpected exam status '{exam.status}'. Only 'approved' allows marks entry."
        )

    # IMP: Enforce RBAC
    verify_marks_entry_access(db, current_user, exam)

    # Pre-fetch offering details for cache invalidation (Sync DB access)
    offering_id = exam.offering_id
    program_id = None
    if offering_id:
        from app.models.subject_offering import SubjectOffering
        off = db.query(SubjectOffering).filter(SubjectOffering.id == offering_id).first()
        if off:
            program_id = off.program_id
    
    saved_count = 0
    audit_entries = []
    
    # Pre-fetch valid students (USNs)
    # Front-end sends student_id as USN (string).
    usns = [entry.student_id for entry in marks_data.marks]
    valid_students = db.query(Student.usn).filter(
        Student.usn.in_(usns),
        # Ensure student belongs to the exam's cohort (Security check)
        Student.cohort_id == exam.cohort_id 
    ).all()
    valid_usns = {s.usn for s in valid_students}
    
    for entry in marks_data.marks:
        if entry.student_id not in valid_usns:
            continue # Skip invalid students (or from wrong cohort)
            
        usn = entry.student_id
        
        # [Validation] Check Max Marks
        sub_q = db.query(SubQuestion).filter(SubQuestion.id == entry.sub_question_id).first()
        if not sub_q:
            continue # Skip invalid sub-question (or raise error)
            
        if entry.marks < 0 or entry.marks > sub_q.max_marks:
            raise HTTPException(
                status_code=400,
                detail=f"Marks {entry.marks} out of range [0, {sub_q.max_marks}] for sub-question {sub_q.label}"
            )

        # Check if mark already exists
        existing = db.query(StudentQuestionMark).filter(
            StudentQuestionMark.exam_id == exam_id,
            StudentQuestionMark.usn == usn,
            StudentQuestionMark.sub_question_id == entry.sub_question_id
        ).first()
        
        if existing:
            old_value = float(existing.marks or 0)
            existing.marks = Decimal(str(entry.marks))
            existing.entered_by = current_user.user_id
            existing.entered_at = datetime.utcnow()
            
            # Audit log for edit
            audit_entries.append({
                "action": "MARKS_EDIT",
                "entity_type": "student_question_marks",
                "entity_id": str(existing.id),
                "old_value": str(old_value),
                "new_value": str(entry.marks),
                "student_id": usn, # Logging USN now
                "sub_question_id": str(entry.sub_question_id)
            })
        else:
            new_mark = StudentQuestionMark(
                id=uuid_lib.uuid4(),
                exam_id=exam_id,
                usn=usn, # USN IS KING
                sub_question_id=entry.sub_question_id,
                marks=Decimal(str(entry.marks)),
                entered_by=current_user.user_id
            )
            db.add(new_mark)
            
            # Audit log for new entry
            audit_entries.append({
                "action": "MARKS_ENTRY",
                "entity_type": "student_question_marks",
                "entity_id": str(new_mark.id),
                "old_value": None,
                "new_value": str(entry.marks),
                "student_id": usn,
                "sub_question_id": str(entry.sub_question_id)
            })
        saved_count += 1
    
    try:
        # Write audit logs
        for audit_data in audit_entries:
            audit_log = AuditLog(
                id=uuid_lib.uuid4(),
                user_id=current_user.user_id,
                action=audit_data["action"],
                entity_type=audit_data["entity_type"],
                entity_id=audit_data["entity_id"],
                old_value=audit_data["old_value"],
                new_value=audit_data["new_value"],
                reason=f"Marks for student {audit_data['student_id']}, subq {audit_data['sub_question_id']}"
            )
            db.add(audit_log)
        
        db.commit()
        
        # Invalidate Cache (Background)
        if offering_id:
            background_tasks.add_task(invalidate_analytics_cache_bg, offering_id, program_id)
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")
        
    return {
        "success": True,
        "success_count": saved_count, # Assuming saved_count is the success count
        "error_count": len(marks_data.marks) - saved_count, # Assuming errors are skipped entries
        "errors": [] # No specific error details are captured in the current loop
    }


@router.get("/compute/{exam_id}")
def compute_marks(
    exam_id: UUID,
    db: Session = Depends(get_db),
    current_user: Profile = Depends(require_teacher_or_above)
):
    """
    Compute marks for an exam on-demand.
    
    NOTE: This endpoint now computes marks dynamically without storing them.
    Per the "no stored analytics" rule, all derived values are computed on-demand.
    For detailed analytics, use the Phase-2B Analytics APIs.
    """
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    # [RBAC] Verify access
    verify_marks_entry_access(db, current_user, exam)
    
    # Get all sections with their selection modes
    sections = db.query(ExamSection).filter(ExamSection.exam_id == exam_id).all()
    
    # Get all student marks (USN-based)
    all_marks = db.query(StudentQuestionMark).filter(StudentQuestionMark.exam_id == exam_id).all()
    
    # Group marks by student (USN)
    student_marks_map = {}
    for mark in all_marks:
        if mark.usn not in student_marks_map:
            student_marks_map[mark.usn] = []
        student_marks_map[mark.usn].append(mark)
    
    results = []
    
    # Get map of USN -> Student UUID for response
    all_usns = list(student_marks_map.keys())
    students = db.query(Student).filter(Student.usn.in_(all_usns)).all()
    usn_to_uuid = {s.usn: str(s.user_id) for s in students}

    for usn, marks in student_marks_map.items():
        total = Decimal(0)
        selected_questions = []
        
        for section in sections:
            section_questions = db.query(Question).filter(Question.section_id == section.id).all()
            
            section_total = Decimal(0)
            
            if section.selection_mode == "BEST_N":
                # Calculate total for each question
                question_totals = []
                for question in section_questions:
                    sub_questions = db.query(SubQuestion).filter(SubQuestion.question_id == question.id).all()
                    sq_ids = [sq.id for sq in sub_questions]
                    
                    q_marks = sum(
                        m.marks for m in marks 
                        if m.sub_question_id in sq_ids
                    )
                    question_totals.append((question.id, q_marks))
                
                # Sort and take best N
                question_totals.sort(key=lambda x: x[1], reverse=True)
                best_questions = question_totals[:section.required_questions]
                
                for q_id, q_total in best_questions:
                    section_total += q_total
                    selected_questions.append(str(q_id))
            else:
                # FIRST_N - sum all marks
                for question in section_questions[:section.required_questions]:
                    sub_questions = db.query(SubQuestion).filter(SubQuestion.question_id == question.id).all()
                    sq_ids = [sq.id for sq in sub_questions]
                    
                    q_marks = sum(
                        m.marks for m in marks 
                        if m.sub_question_id in sq_ids
                    )
                    section_total += q_marks
                    selected_questions.append(str(question.id))
            
            # [FIX] Enforce Section Max Marks Cap
            # If student scores 12 in a 10-mark section (via extras), cap it at 10.
            if section_total > section.max_marks:
                section_total = Decimal(section.max_marks)
                
            total += section_total
        
        results.append({
            "student_id": usn_to_uuid.get(usn, "UNKNOWN"), # Return UUID if possible
            "student_usn": usn, # Add USN for clarity
            "exam_id": str(exam_id),
            "total_marks": float(total),
            "selected_questions": selected_questions,
            "computed_at": datetime.utcnow().isoformat(),
            "note": "Computed on-demand, not stored"
        })
    
    return {
        "exam_id": str(exam_id),
        "student_count": len(results),
        "results": results
    }


@router.get("/student/{student_id}", response_model=List[StudentMarksResponse])
def get_student_marks(
    student_id: UUID,
    db: Session = Depends(get_db),
    current_user: Profile = Depends(require_authenticated)
):
    """Get marks for a specific student (students can only view their own)."""
    # Check authorization
    from app.models import UserRole
    role = db.query(UserRole).filter(UserRole.user_id == current_user.user_id).first()
    user_role = role.role.value if role else "student"
    
    if user_role == "student" and current_user.user_id != student_id:
        raise HTTPException(status_code=403, detail="Not authorized to view other student's marks")
    
    # Updated to use USN for StudentQuestionMark
    # First get student USN from UUID
    student = db.query(Student).filter(Student.user_id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
        
    marks = db.query(StudentQuestionMark).filter(StudentQuestionMark.usn == student.usn).all()
    
    # Map back to response format
    return [
        {
            "id": m.id,
            "exam_id": m.exam_id,
            "student_id": student.usn, # Return USN (schema expects str)
            "sub_question_id": m.sub_question_id,
            "marks": m.marks,
            "entered_at": m.entered_at
        }
        for m in marks
    ]


# ============= APPROVAL WORKFLOW =============

@router.post("/submit-for-approval/{exam_id}")
def submit_for_approval(
    exam_id: UUID,
    db: Session = Depends(get_db),
    current_user: Profile = Depends(require_teacher_or_above)
):
    """
    Submit exam marks for HOD approval.
    Faculty can submit, only HOD/Principal can approve.
    RBAC: Teacher (submitter).
    """
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    # [RBAC] Verify access
    verify_marks_entry_access(db, current_user, exam)
    
    if exam.status not in ["draft", "rejected"]:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot submit exam with status '{exam.status}'"
        )
    
    # Verify marks exist (USN-based)
    marks_count = db.query(StudentQuestionMark).filter(StudentQuestionMark.exam_id == exam_id).count()
    if marks_count == 0:
        raise HTTPException(status_code=400, detail="No marks entered for this exam")
    
    exam.status = "submitted"
    exam.submitted_at = datetime.utcnow()
    
    # Audit Log
    from app.models.audit import AuditLog
    audit = AuditLog(
        id=uuid_lib.uuid4(),
        table_name="exams",
        record_id=exam_id,
        action="SUBMIT",
        old_data={"status": "draft"},
        new_data={"status": "submitted"},
        user_id=current_user.user_id,
        reason="Exam submitted for approval"
    )
    db.add(audit)
    
    db.commit()
    
    return {
        "success": True,
        "exam_id": str(exam_id),
        "status": "submitted",
        "submitted_by": str(current_user.user_id),
        "submitted_at": exam.submitted_at.isoformat()
    }


@router.post("/approve/{exam_id}")
def approve_marks(
    exam_id: UUID,
    db: Session = Depends(get_db),
    current_user: Profile = Depends(require_teacher_or_above)
):
    """
    Approve submitted exam marks.
    Only HOD/Principal can approve marks.
    RBAC: HOD/Principal (approver).
    """
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    # [RBAC] Verify access (HOD=Dept logic included here)
    verify_marks_entry_access(db, current_user, exam)
    
    # Verify HOD or above for approval
    from app.models import UserRole
    from app.core.permissions import AppRole
    role_record = db.query(UserRole).filter(UserRole.user_id == current_user.user_id).first()
    if not role_record or role_record.role not in [AppRole.HOD, AppRole.PRINCIPAL]:
        raise HTTPException(status_code=403, detail="Only HOD or Principal can approve marks")
    
    if exam.status != "submitted":
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot approve exam with status '{exam.status}'. Must be 'submitted'."
        )
    
    exam.approved_at = datetime.utcnow()
    exam.approved_by = current_user.user_id
    
    # Audit Log
    from app.models.audit import AuditLog
    audit = AuditLog(
        id=uuid_lib.uuid4(),
        table_name="exams",
        record_id=exam_id,
        action="APPROVE",
        old_data={"status": "submitted"},
        new_data={"status": "approved"},
        user_id=current_user.user_id,
        reason="Exam marks approved"
    )
    db.add(audit)
    
    db.commit()
    
    return {
        "success": True,
        "exam_id": str(exam_id),
        "status": "approved",
        "approved_by": str(current_user.user_id),
        "approved_at": exam.approved_at.isoformat()
    }


from pydantic import BaseModel as PydanticBaseModel

class RejectRequest(PydanticBaseModel):
    reason: str


@router.post("/reject/{exam_id}")
def reject_marks(
    exam_id: UUID,
    data: RejectRequest,
    db: Session = Depends(get_db),
    current_user: Profile = Depends(require_teacher_or_above)
):
    """
    Reject submitted exam marks with reason.
    Only HOD/Principal can reject marks.
    RBAC: HOD/Principal (rejector).
    """
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
        
    # [RBAC] Verify access (HOD=Dept logic included here)
    verify_marks_entry_access(db, current_user, exam)
    
    # Verify HOD or above for rejection
    from app.models import UserRole
    from app.core.permissions import AppRole
    role_record = db.query(UserRole).filter(UserRole.user_id == current_user.user_id).first()
    if not role_record or role_record.role not in [AppRole.HOD, AppRole.PRINCIPAL]:
        raise HTTPException(status_code=403, detail="Only HOD or Principal can reject marks")
    
    if exam.status != "submitted":
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot reject exam with status '{exam.status}'. Must be 'submitted'."
        )
    
    exam.status = "rejected"
    db.commit()
    
    # Audit log the rejection with reason
    from app.models.audit import AuditLog
    audit = AuditLog(
        id=uuid_lib.uuid4(),
        table_name="exams",
        record_id=exam_id,
        action="REJECT",
        old_data=None,
        new_data={"status": "rejected", "reason": data.reason},
        user_id=current_user.user_id
    )
    db.add(audit)
    db.commit()
    
    return {
        "success": True,
        "exam_id": str(exam_id),
        "status": "rejected",
        "rejected_by": str(current_user.user_id),
        "reason": data.reason
    }


@router.post("/lock/{exam_id}")
def lock_marks(
    exam_id: UUID,
    db: Session = Depends(get_db),
    current_user: Profile = Depends(require_teacher_or_above)
):
    """
    Lock approved marks - no further edits allowed.
    Only HOD/Principal can lock after approval.
    RBAC: HOD/Principal.
    """
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
        
    # [RBAC] Verify access (HOD=Dept logic included here)
    verify_marks_entry_access(db, current_user, exam)
    
    # Verify HOD or above for locking
    from app.models import UserRole
    from app.core.permissions import AppRole
    role_record = db.query(UserRole).filter(UserRole.user_id == current_user.user_id).first()
    if not role_record or role_record.role not in [AppRole.HOD, AppRole.PRINCIPAL]:
        raise HTTPException(status_code=403, detail="Only HOD or Principal can lock marks")
    
    if exam.status != "approved":
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot lock exam with status '{exam.status}'. Must be 'approved' first."
        )
    
    exam.status = "locked"
    
    # Audit Log
    from app.models.audit import AuditLog
    audit = AuditLog(
        id=uuid_lib.uuid4(),
        table_name="exams",
        record_id=exam_id,
        action="LOCK",
        old_data={"status": "approved"},
        new_data={"status": "locked"},
        user_id=current_user.user_id,
        reason="Exam locked (immutable)"
    )
    db.add(audit)
    
    db.commit()
    
    return {
        "success": True,
        "exam_id": str(exam_id),
        "status": "locked",
        "locked_by": str(current_user.user_id)
    }


# ============= BULK IMPORT/EXPORT =============

from fastapi import UploadFile, File
from fastapi.responses import StreamingResponse
import csv
import io


@router.get("/template/{exam_id}")
def get_marks_template(
    exam_id: UUID,
    db: Session = Depends(get_db),
    current_user: Profile = Depends(require_teacher_or_above)
):
    """
    Download marks entry template for an exam.
    Template includes: USN, Student Name, and columns for each sub-question.
    """
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    # [RBAC] Verify access
    verify_marks_entry_access(db, current_user, exam)
    
    # Get all sections, questions, sub-questions
    sections = db.query(ExamSection).filter(ExamSection.exam_id == exam_id).order_by(ExamSection.sequence).all()
    
    # Build header row
    headers = ["USN", "Student Name"]
    sq_mapping = []  # (sub_question_id, column_name, max_marks)
    
    for section in sections:
        questions = db.query(Question).filter(Question.section_id == section.id).order_by(Question.sequence).all()
        for question in questions:
            sub_questions = db.query(SubQuestion).filter(SubQuestion.question_id == question.id).all()
            for sq in sub_questions:
                col_name = f"S{section.name}_Q{question.sequence}_{sq.label or 'a'} (Max:{sq.max_marks})"
                headers.append(col_name)
                sq_mapping.append((str(sq.id), col_name, float(sq.max_marks)))
    
    # Get students for this cohort
    students = db.query(Student).filter(
        Student.cohort_id == exam.cohort_id,
        Student.status == "active"
    ).order_by(Student.usn).all()

    # [NEW] Fetch existing marks to pre-fill
    existing_marks = db.query(StudentQuestionMark).filter(
        StudentQuestionMark.exam_id == exam_id
    ).all()
    
    # Map (usn, sq_id) -> marks
    marks_map = {}
    for mark in existing_marks:
        marks_map[(mark.usn, str(mark.sub_question_id))] = float(mark.marks or 0)
    
    # [v2] Create XLSX
    from openpyxl import Workbook
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Marks Template"
    
    # Header
    ws.append(headers)
    
    # Rows
    for student in students:
        row = [student.usn, student.name]
        for sq_id, _, _ in sq_mapping:
            # Pre-fill mark if exists, else empty
            val = marks_map.get((student.usn, sq_id), "")
            row.append(val)
        ws.append(row)
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Return as downloadable file
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=marks_template_{exam_id}.xlsx"
        }
    )


@router.post("/import/{exam_id}")
def import_marks(
    exam_id: UUID,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Profile = Depends(require_teacher_or_above)
):
    """
    Import marks from Excel (.xlsx) or CSV file for an exam.
    
    Expected format matches the XLSX template from /api/v1/export/assessment/marks-template/{exam_id}.
    Headers: USN, Name, {Section}_Q{Num}{Label}_Max{Val}
    """
    from app.models import AuditLog
    import uuid as uuid_lib
    from io import BytesIO
    
    # 1. Basic Checks
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
        
    if exam.status == "locked":
        raise HTTPException(status_code=400, detail="Exam is locked. Cannot import marks.")
    if exam.status != "approved":
        raise HTTPException(
            status_code=400,
            detail=f"Exam status is '{exam.status}'. Marks entry only allowed after HOD approval."
        )

    # 2. RBAC
    verify_marks_entry_access(db, current_user, exam)

    # 3. Read File Content
    content = file.file.read()
    rows = []
    
    if file.filename.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=BytesIO(content), data_only=True)
            ws = wb.active
            # Convert worksheet to list of dicts
            headers = [str(cell.value) for cell in ws[1] if cell.value]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue # Skip empty rows
                rows.append(dict(zip(headers, row)))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to read Excel: {str(e)}")
    elif file.filename.endswith('.csv'):
        try:
            import csv
            from io import StringIO
            decoded = content.decode('utf-8')
            csv_reader = csv.DictReader(StringIO(decoded))
            rows = list(csv_reader)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to read CSV: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="Only .xlsx and .csv files are supported.")

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty or contains no data rows.")

    # 4. Map Sub-Questions
    # Header format: {Section}_Q{Num}{Label}_Max{Val}
    sections = db.query(ExamSection).filter(ExamSection.exam_id == exam_id).all()
    sq_map = {} # {header_name: sq_id}
    sq_max = {} # {sq_id: max_marks}
    
    for section in sections:
        questions = db.query(Question).filter(Question.section_id == section.id).all()
        for q in questions:
            sub_qs = db.query(SubQuestion).filter(SubQuestion.question_id == q.id).all()
            for sq in sub_qs:
                # MATCH TEMPLATE GEN: f"S{section.name}_Q{question.sequence}_{sq.label or 'a'} (Max:{sq.max_marks})"
                col_name = f"S{section.name}_Q{q.sequence}_{sq.label or 'a'} (Max:{sq.max_marks})"
                sq_map[col_name] = sq.id
                sq_max[sq.id] = float(sq.max_marks)

    # 5. Process Rows
    results = {"success_count": 0, "error_count": 0, "errors": []}
    marks_to_add = []
    audit_entries = []
    
    for idx, row in enumerate(rows, 2):
        usn = str(row.get("USN") or "").strip()
        if not usn or usn == "None":
            continue
            
        student = db.query(Student).filter(Student.usn == usn, Student.cohort_id == exam.cohort_id).first()
        if not student:
            results["errors"].append(f"Row {idx}: USN {usn} not found in this cohort.")
            results["error_count"] += 1
            continue

        for col, sq_id in sq_map.items():
            val = row.get(col)
            if val is None or str(val).strip() == "":
                continue
            
            try:
                marks_val = float(val)
            except (ValueError, TypeError):
                results["errors"].append(f"Row {idx}: Invalid marks '{val}' for {col}")
                results["error_count"] += 1
                continue
                
            if marks_val < 0 or marks_val > sq_max[sq_id]:
                results["errors"].append(f"Row {idx}: Marks {marks_val} out of range [0, {sq_max[sq_id]}] for {col}")
                results["error_count"] += 1
                continue
            
            # Check existing
            existing = db.query(StudentQuestionMark).filter(
                StudentQuestionMark.exam_id == exam_id,
                StudentQuestionMark.usn == usn,
                StudentQuestionMark.sub_question_id == sq_id
            ).first()
            
            if existing:
                old_val = float(existing.marks or 0)
                existing.marks = Decimal(str(marks_val))
                existing.entered_by = current_user.user_id
                existing.entered_at = datetime.utcnow()
                
                audit_entries.append({
                    "action": "MARKS_IMPORT_EDIT",
                    "entity_id": str(existing.id),
                    "old_value": str(old_val),
                    "new_value": str(marks_val),
                    "student_id": usn,
                    "sub_question_id": str(sq_id)
                })
            else:
                new_mark = StudentQuestionMark(
                    id=uuid_lib.uuid4(),
                    exam_id=exam_id,
                    usn=usn,
                    sub_question_id=sq_id,
                    marks=Decimal(str(marks_val)),
                    entered_by=current_user.user_id
                )
                db.add(new_mark)
                
                audit_entries.append({
                    "action": "MARKS_IMPORT_NEW",
                    "entity_id": str(new_mark.id),
                    "old_value": None,
                    "new_value": str(marks_val),
                    "student_id": usn,
                    "sub_question_id": str(sq_id)
                })
            results["success_count"] += 1

    # 6. Commit & Audit
    try:
        for a in audit_entries:
            db.add(AuditLog(
                id=uuid_lib.uuid4(),
                user_id=current_user.user_id,
                action=a["action"],
                entity_type="student_question_marks",
                entity_id=a["entity_id"],
                old_value=a["old_value"],
                new_value=a["new_value"],
                reason=f"Excel/CSV Import - Student {a['student_id']}"
            ))
        db.commit()
        
        # Invalidate Cache
        if exam.offering_id:
            background_tasks.add_task(invalidate_analytics_cache_bg, exam.offering_id, None)
            
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error during import: {str(e)}")

    return {
        "success": True,
        "exam_id": str(exam_id),
        "imported_marks": results["success_count"],
        "errors": results["errors"][:10],
        "total_errors": results["error_count"]
    }
